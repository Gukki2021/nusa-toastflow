#!/usr/bin/env python3
"""
Extract the recommendation engine's data from the club's master spreadsheet.

Usage:
    pip install openpyxl
    python scripts/extract-data.py "NUSA TM 25-26 Program Plan & Distribution List.xlsx"

Writes (all git-ignored except projects.json):
    data/members.json   – roster: name, Pathways credentials, email
    data/history.json   – every past role appointment (fuel for recommendations)
    data/projects.json  – Pathways project objectives catalogue

Run this again whenever the master sheet is updated so recommendations stay current.
"""
import sys, json, re, datetime, os
import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "data")


def norm_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", v.strip())
        if m:
            return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    members = []
    for row in wb["Registered Members"].iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            members.append({
                "name": str(row[0]).strip(),
                "credentials": str(row[1]).strip() if row[1] else "",
                "email": str(row[2]).strip() if row[2] else "",
            })

    ws = wb["Meeting Appointment Holders"]
    rows = list(ws.iter_rows(values_only=True))
    dates = [norm_date(c) for c in rows[0][1:]]
    themes = [str(c).strip() if c else "" for c in rows[1][1:]]
    history = []
    for r in rows[2:]:
        if not r[0] or not str(r[0]).strip():
            continue
        role = str(r[0]).strip()
        for j, cell in enumerate(r[1:]):
            if j >= len(dates) or not dates[j]:
                continue
            if cell and str(cell).strip() and str(cell).strip().lower() != "x":
                history.append({
                    "date": dates[j],
                    "theme": themes[j] if j < len(themes) else "",
                    "role": role,
                    "raw": str(cell).strip(),
                })

    projects = []
    for row in wb["Project Objectives"].iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            projects.append({
                "level": str(row[0]).strip(),
                "project": str(row[1]).strip() if row[1] else "",
                "objective": str(row[2]).strip() if row[2] else "",
                "timing": str(row[3]).strip() if row[3] else "",
            })

    os.makedirs(OUT, exist_ok=True)
    for name, data in [("members", members), ("history", history), ("projects", projects)]:
        with open(os.path.join(OUT, f"{name}.json"), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=1, ensure_ascii=False)
        print(f"  data/{name}.json  ({len(data)} rows)")
    print("Done. Note: members.json and history.json are git-ignored (contain member data).")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/extract-data.py <master-spreadsheet.xlsx>")
    main(sys.argv[1])
